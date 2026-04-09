#!/usr/bin/env python3
"""
Valida registro a registro el archivo NUEVABASE.xlsx vs tabla public.prestamos.

Uso:
  export DATABASE_URL='postgresql://...'
  export EXCEL_PATH='/Users/paulgarcia/Desktop/NUEVABASE.xlsx'
  python3 validate_nuevabase_vs_db.py
"""

import csv
import os
import re
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

import psycopg2
from openpyxl import load_workbook


DATABASE_URL = os.environ.get("DATABASE_URL")
EXCEL_PATH = os.environ.get("EXCEL_PATH", "/Users/paulgarcia/Desktop/NUEVABASE.xlsx")
OUT_CSV = os.environ.get("OUT_CSV", "reporte_validacion_nuevabase_vs_db.csv")


def clean_text(value):
    return " ".join(str(value or "").strip().split())


def key_name(value):
    return clean_text(value).upper()


def to_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None:
        return None
    text = clean_text(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d-%b-%y", "%d-%b-%Y", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def to_decimal(value, default="0"):
    if value is None:
        return Decimal(default)
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    text = clean_text(value).replace(",", "").replace("$", "")
    if text in {"", "-", "N/A"}:
        return Decimal(default)
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal(default)


def to_int(value, default=0):
    return int(to_decimal(value, str(default)))


def parse_status_pending(status_text):
    text = key_name(status_text)
    match = re.search(r"LE\s+QUEDAN\s+(\d+)\s+PAGOS?", text)
    if match:
        return int(match.group(1))
    if text in {"NO DEBE NADA", "PAGADO"}:
        return 0
    return None


def load_excel_rows():
    ws = load_workbook(EXCEL_PATH, data_only=True)["Sheet1"]
    data = []
    sequence = defaultdict(int)

    for row_num in range(2, ws.max_row + 1):
        nombre = clean_text(ws.cell(row_num, 4).value)
        if not nombre:
            continue
        fecha_inicio = to_date(ws.cell(row_num, 1).value)
        if not fecha_inicio:
            continue
        monto_solicitado = to_decimal(ws.cell(row_num, 5).value)
        base_key = (key_name(nombre), fecha_inicio.isoformat(), str(monto_solicitado.quantize(Decimal("0.01"))))
        sequence[base_key] += 1
        seq = sequence[base_key]

        status = clean_text(ws.cell(row_num, 30).value)
        data.append(
            {
                "excel_row": row_num,
                "nombre": nombre,
                "fecha_inicio": fecha_inicio,
                "monto_solicitado": monto_solicitado.quantize(Decimal("0.01")),
                "seq": seq,
                "pagos_hechos": to_int(ws.cell(row_num, 26).value, 0),
                "pagos_pendientes": to_int(ws.cell(row_num, 27).value, 0),
                "pagado": to_decimal(ws.cell(row_num, 28).value).quantize(Decimal("0.01")),
                "pendiente": to_decimal(ws.cell(row_num, 29).value).quantize(Decimal("0.01")),
                "status": status,
                "status_pending": parse_status_pending(status),
            }
        )
    return data


def load_db_rows(conn):
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT
              id,
              nombre_completo,
              fecha_inicio::date,
              monto_solicitado,
              pagos_hechos,
              pagos_pendientes,
              pagado,
              pendiente,
              status
            FROM public.prestamos
            ORDER BY fecha_inicio, id
            """
        )
        rows = cur.fetchall()

    grouped = defaultdict(list)
    for item in rows:
        (
            prestamo_id,
            nombre,
            fecha_inicio,
            monto_solicitado,
            pagos_hechos,
            pagos_pendientes,
            pagado,
            pendiente,
            status,
        ) = item
        monto = to_decimal(monto_solicitado).quantize(Decimal("0.01"))
        key = (key_name(nombre), fecha_inicio.isoformat(), str(monto))
        grouped[key].append(
            {
                "id": str(prestamo_id),
                "nombre": clean_text(nombre),
                "fecha_inicio": fecha_inicio,
                "monto_solicitado": monto,
                "pagos_hechos": int(pagos_hechos or 0),
                "pagos_pendientes": int(pagos_pendientes or 0),
                "pagado": to_decimal(pagado).quantize(Decimal("0.01")),
                "pendiente": to_decimal(pendiente).quantize(Decimal("0.01")),
                "status": clean_text(status),
            }
        )
    return grouped


def compare_rows(excel_rows, db_grouped):
    result = []
    grouped_seq = defaultdict(int)

    for excel in excel_rows:
        k = (
            key_name(excel["nombre"]),
            excel["fecha_inicio"].isoformat(),
            str(excel["monto_solicitado"]),
        )
        grouped_seq[k] += 1
        seq = grouped_seq[k]
        candidates = db_grouped.get(k, [])
        db = candidates[seq - 1] if len(candidates) >= seq else None

        if not db:
            result.append(
                {
                    "excel_row": excel["excel_row"],
                    "tipo": "NO_MATCH_DB",
                    "detalle": "No existe préstamo equivalente en DB",
                    "nombre": excel["nombre"],
                    "fecha_inicio": excel["fecha_inicio"].isoformat(),
                    "monto_solicitado": str(excel["monto_solicitado"]),
                    "prestamo_id": "",
                }
            )
            continue

        diffs = []
        for field in ["pagos_hechos", "pagos_pendientes", "pagado", "pendiente", "status"]:
            if str(excel[field]).strip().upper() != str(db[field]).strip().upper():
                diffs.append(f"{field}: excel={excel[field]} db={db[field]}")

        if excel["status_pending"] is not None and excel["status_pending"] != db["pagos_pendientes"]:
            diffs.append(
                f"status_pending: excel={excel['status_pending']} db={db['pagos_pendientes']}"
            )

        if diffs:
            result.append(
                {
                    "excel_row": excel["excel_row"],
                    "tipo": "DIFERENCIAS",
                    "detalle": " | ".join(diffs),
                    "nombre": excel["nombre"],
                    "fecha_inicio": excel["fecha_inicio"].isoformat(),
                    "monto_solicitado": str(excel["monto_solicitado"]),
                    "prestamo_id": db["id"],
                }
            )

    return result


def main():
    if not DATABASE_URL:
        raise RuntimeError("Falta DATABASE_URL en variables de entorno.")

    excel_rows = load_excel_rows()
    conn = psycopg2.connect(DATABASE_URL)
    db_grouped = load_db_rows(conn)
    conn.close()

    diffs = compare_rows(excel_rows, db_grouped)

    with open(OUT_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "excel_row",
                "tipo",
                "detalle",
                "nombre",
                "fecha_inicio",
                "monto_solicitado",
                "prestamo_id",
            ],
        )
        writer.writeheader()
        writer.writerows(diffs)

    print(f"Excel filas analizadas: {len(excel_rows)}")
    print(f"Diferencias encontradas: {len(diffs)}")
    print(f"Reporte: {OUT_CSV}")


if __name__ == "__main__":
    main()
