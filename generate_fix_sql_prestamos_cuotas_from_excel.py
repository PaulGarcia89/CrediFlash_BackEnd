#!/usr/bin/env python3
import datetime as dt
import os
import re
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP

import openpyxl

EXCEL_PATH = os.getenv("EXCEL_PATH", "/Users/paulgarcia/Desktop/cargaAbril13.xlsx")
OUTPUT_SQL = os.getenv("OUTPUT_SQL", "fix_prestamos_cuotas_from_cargaAbril13.sql")
OUTPUT_REPORT = os.getenv("OUTPUT_REPORT", "reporte_prestamos_parse_cargaAbril13.csv")
SHEET_CANDIDATES = ["CONTROL PRESTAMOS", "Sheet1", "SHEET1"]


def clean_text(value):
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def upper(value):
    return clean_text(value).upper()


def to_decimal(value, default=Decimal("0.00")):
    if value is None:
        return default
    if isinstance(value, (int, float, Decimal)):
        try:
            return Decimal(str(value))
        except InvalidOperation:
            return default

    txt = clean_text(value)
    if txt in {"", "-", "None", "null"}:
        return default

    txt = txt.replace(",", "").replace("$", "").replace("%", "")
    try:
        return Decimal(txt)
    except InvalidOperation:
        return default


def money(value):
    return to_decimal(value).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def to_int(value, default=0):
    try:
        return int(to_decimal(value, Decimal(default)))
    except Exception:
        return default


def to_date(value):
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value

    txt = clean_text(value)
    if not txt or txt == "-":
        return None

    for fmt in ("%d-%b-%y", "%d-%b-%Y", "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return dt.datetime.strptime(txt, fmt).date()
        except ValueError:
            continue

    return None


def parse_interes_to_int(value):
    x = to_decimal(value, Decimal("0"))
    if x <= 1:
        x = (x * Decimal("100")).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
    return int(x)


def parse_tasa_variable(value):
    x = to_decimal(value, Decimal("0.12"))
    if x > 1:
        x = (x / Decimal("100")).quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)
    if x < 0:
        x = Decimal("0.12")
    return x.quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)


def normalize_modalidad(value):
    v = upper(value)
    if v in {"SEMANAL", "QUINCENAL", "MENSUAL"}:
        return v
    return "SEMANAL"


def extract_pending_from_status(value):
    txt = upper(value)
    m = re.search(r"LE QUEDAN\s+(\d+)\s+PAGOS", txt)
    if m:
        return int(m.group(1))
    if txt in {"NO DEBE NADA", "PAGADO"}:
        return 0
    return None


def is_checked_cell(value):
    if value is True:
        return True
    if value is False or value is None:
        return False
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value)) > 0

    txt = upper(value).replace(".", "")
    return txt in {
        "X", "SI", "SÍ", "TRUE", "VERDADERO", "1", "✓", "✔", "☑", "☒", "CHECK", "CHECKED"
    }


def esc_sql(text):
    return str(text).replace("'", "''")


def pick_sheet(workbook):
    available = {upper(name): name for name in workbook.sheetnames}
    for candidate in SHEET_CANDIDATES:
        key = upper(candidate)
        if key in available:
            return workbook[available[key]]
    raise KeyError(f"No se encontró hoja de préstamos. Disponibles: {workbook.sheetnames}")


def parse_rows():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = pick_sheet(wb)

    rows = []
    seq_by_key = {}

    for r in range(4, ws.max_row + 1):
        nombre = clean_text(ws.cell(r, 4).value)
        if not nombre or upper(nombre) in {"NOMBRE", "TOTAL"}:
            continue

        fecha_inicio = to_date(ws.cell(r, 1).value)
        if not fecha_inicio:
            continue

        key = (upper(nombre), fecha_inicio.isoformat())
        seq_by_key[key] = seq_by_key.get(key, 0) + 1
        seq_excel = seq_by_key[key]

        monto_solicitado = money(ws.cell(r, 5).value)
        interes_int = parse_interes_to_int(ws.cell(r, 6).value)
        tasa_variable = parse_tasa_variable(ws.cell(r, 6).value)
        modalidad = normalize_modalidad(ws.cell(r, 7).value)
        num_semanas = to_int(ws.cell(r, 8).value, 0)
        num_dias = to_int(ws.cell(r, 9).value, 0)
        fecha_vencimiento = to_date(ws.cell(r, 10).value)

        total_pagar = money(ws.cell(r, 11).value)
        ganancias = money(ws.cell(r, 12).value)
        pagos_semanales = money(ws.cell(r, 13).value)

        pagos_hechos_col = to_int(ws.cell(r, 26).value, 0)
        pagos_pend_col = to_int(ws.cell(r, 27).value, 0)
        pagado_col = money(ws.cell(r, 28).value)
        pendiente_col = money(ws.cell(r, 29).value)
        status_col = clean_text(ws.cell(r, 30).value)

        check_values = [ws.cell(r, c).value for c in range(14, 26)]
        pagos_hechos_checks = sum(1 for v in check_values if is_checked_cell(v))

        if num_semanas <= 0 and pagos_semanales > 0:
            num_semanas = int((total_pagar / pagos_semanales).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
        if num_semanas <= 0:
            num_semanas = max(1, pagos_hechos_col, pagos_hechos_checks)

        resumen_informado = any(clean_text(ws.cell(r, c).value) not in {"", "-"} for c in (26, 27, 30))

        if resumen_informado:
            pending_from_status = extract_pending_from_status(status_col)
            if pending_from_status is not None:
                pagos_pendientes = pending_from_status
                pagos_hechos = max(0, num_semanas - pagos_pendientes)
            else:
                pagos_hechos = max(0, pagos_hechos_col)
                pagos_pendientes = pagos_pend_col if pagos_pend_col > 0 else max(0, num_semanas - pagos_hechos)
        else:
            pagos_hechos = max(0, pagos_hechos_checks)
            pagos_pendientes = max(0, num_semanas - pagos_hechos)

        pagos_hechos = min(max(0, pagos_hechos), num_semanas)
        pagos_pendientes = min(max(0, pagos_pendientes), num_semanas)

        if pagos_semanales <= 0 and num_semanas > 0:
            pagos_semanales = (total_pagar / Decimal(num_semanas)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

        if total_pagar <= 0:
            total_pagar = (monto_solicitado + (monto_solicitado * tasa_variable * Decimal(num_semanas))).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

        if ganancias <= 0:
            ganancias = (total_pagar - monto_solicitado).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

        if resumen_informado and pagado_col > 0:
            pagado = pagado_col
        else:
            pagado = (pagos_semanales * Decimal(pagos_hechos)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            pagado = min(pagado, total_pagar)

        if resumen_informado and pendiente_col > 0:
            pendiente = pendiente_col
        else:
            pendiente = max(Decimal("0.00"), (total_pagar - pagado)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

        status = "PAGADO" if pagos_pendientes == 0 or pendiente <= 0 else f"LE QUEDAN {pagos_pendientes} PAGOS POR PAGAR"

        rows.append({
            "excel_row": r,
            "seq_excel": seq_excel,
            "nombre_completo": nombre,
            "fecha_inicio": fecha_inicio.isoformat(),
            "monto_solicitado": monto_solicitado,
            "interes": interes_int,
            "tasa_variable": tasa_variable,
            "modalidad": modalidad,
            "num_semanas": num_semanas,
            "num_dias": num_dias,
            "fecha_vencimiento": fecha_vencimiento.isoformat() if fecha_vencimiento else None,
            "total_pagar": total_pagar,
            "ganancias": ganancias,
            "pagos_semanales": pagos_semanales,
            "pagos_hechos": pagos_hechos,
            "pagos_pendientes": pagos_pendientes,
            "pagado": pagado,
            "pendiente": pendiente,
            "status": status,
        })

    return rows


def build_sql(rows):
    values_lines = []
    for row in rows:
        fecha_vto_sql = f"'{row['fecha_vencimiento']}'::date" if row["fecha_vencimiento"] else "NULL"
        values_lines.append(
            "("
            f"'{esc_sql(row['nombre_completo'])}', "
            f"'{row['fecha_inicio']}'::date, "
            f"{row['seq_excel']}, "
            f"{row['monto_solicitado']}, "
            f"{row['interes']}, "
            f"{row['tasa_variable']}, "
            f"'{esc_sql(row['modalidad'])}', "
            f"{row['num_semanas']}, "
            f"{row['num_dias']}, "
            f"{fecha_vto_sql}, "
            f"{row['total_pagar']}, "
            f"{row['ganancias']}, "
            f"{row['pagos_semanales']}, "
            f"{row['pagos_hechos']}, "
            f"{row['pagos_pendientes']}, "
            f"{row['pagado']}, "
            f"{row['pendiente']}, "
            f"'{esc_sql(row['status'])}'"
            ")"
        )

    values_block = ",\n    ".join(values_lines)

    return f"""-- Generado automáticamente desde {EXCEL_PATH}
BEGIN;

WITH excel_data AS (
  SELECT * FROM (
    VALUES
    {values_block}
  ) AS v(
    nombre_completo,
    fecha_inicio,
    seq_excel,
    monto_solicitado,
    interes,
    tasa_variable,
    modalidad,
    num_semanas,
    num_dias,
    fecha_vencimiento,
    total_pagar,
    ganancias,
    pagos_semanales,
    pagos_hechos,
    pagos_pendientes,
    pagado,
    pendiente,
    status
  )
),
db_rows AS (
  SELECT
    p.id,
    upper(trim(regexp_replace(p.nombre_completo, '\\\\s+', ' ', 'g'))) AS nombre_norm,
    p.fecha_inicio::date AS fecha_inicio,
    row_number() OVER (
      PARTITION BY upper(trim(regexp_replace(p.nombre_completo, '\\\\s+', ' ', 'g'))), p.fecha_inicio::date
      ORDER BY p.id
    ) AS seq_db
  FROM public.prestamos p
),
xls_norm AS (
  SELECT
    upper(trim(regexp_replace(e.nombre_completo, '\\\\s+', ' ', 'g'))) AS nombre_norm,
    e.*
  FROM excel_data e
),
matches AS (
  SELECT
    d.id AS prestamo_id,
    x.*
  FROM db_rows d
  JOIN xls_norm x
    ON x.nombre_norm = d.nombre_norm
   AND x.fecha_inicio = d.fecha_inicio
   AND x.seq_excel = d.seq_db
),
updated AS (
  UPDATE public.prestamos p
  SET
    monto_solicitado = m.monto_solicitado,
    interes = m.interes,
    modalidad = m.modalidad,
    num_semanas = m.num_semanas,
    num_dias = m.num_dias,
    fecha_vencimiento = m.fecha_vencimiento,
    total_pagar = m.total_pagar,
    ganancias = m.ganancias,
    pagos_semanales = m.pagos_semanales,
    pagos_hechos = m.pagos_hechos,
    pagos_pendientes = m.pagos_pendientes,
    pagado = m.pagado,
    pendiente = m.pendiente,
    status = m.status,
    fecha_aprobacion = COALESCE(p.fecha_aprobacion, p.fecha_inicio)
  FROM matches m
  WHERE p.id = m.prestamo_id
  RETURNING p.id
),
deleted AS (
  DELETE FROM public.cuotas c
  USING updated u
  WHERE c.prestamo_id = u.id
  RETURNING c.id
),
seed AS (
  SELECT
    p.id AS prestamo_id,
    p.fecha_inicio::date AS fecha_inicio,
    GREATEST(COALESCE(p.num_semanas,0), 1)::int AS num_semanas,
    GREATEST(COALESCE(p.pagos_hechos,0), 0)::int AS pagos_hechos,
    COALESCE(p.monto_solicitado,0)::numeric(15,2) AS monto_solicitado,
    COALESCE(p.ganancias,0)::numeric(15,2) AS ganancias,
    COALESCE(p.total_pagar,0)::numeric(15,2) AS total_pagar
  FROM public.prestamos p
  JOIN updated u ON u.id = p.id
),
series AS (
  SELECT s.*, gs.n AS cuota_num
  FROM seed s
  CROSS JOIN LATERAL generate_series(1, s.num_semanas) AS gs(n)
),
calc AS (
  SELECT
    prestamo_id,
    cuota_num,
    (fecha_inicio + (cuota_num * 7) * INTERVAL '1 day')::date AS fecha_vencimiento,
    ROUND((monto_solicitado / num_semanas)::numeric, 2) AS cap_base,
    ROUND((ganancias / num_semanas)::numeric, 2) AS int_base,
    ROUND((total_pagar / num_semanas)::numeric, 2) AS tot_base,
    monto_solicitado,
    ganancias,
    total_pagar,
    num_semanas,
    pagos_hechos
  FROM series
),
ins AS (
  INSERT INTO public.cuotas (
    prestamo_id,
    fecha_vencimiento,
    monto_capital,
    monto_interes,
    monto_total,
    estado,
    fecha_pago,
    monto_pagado,
    observaciones,
    created_at
  )
  SELECT
    c.prestamo_id,
    c.fecha_vencimiento,
    CASE
      WHEN c.cuota_num < c.num_semanas THEN c.cap_base
      ELSE ROUND((c.monto_solicitado - (c.cap_base * (c.num_semanas - 1)))::numeric, 2)
    END,
    CASE
      WHEN c.cuota_num < c.num_semanas THEN c.int_base
      ELSE ROUND((c.ganancias - (c.int_base * (c.num_semanas - 1)))::numeric, 2)
    END,
    CASE
      WHEN c.cuota_num < c.num_semanas THEN c.tot_base
      ELSE ROUND((c.total_pagar - (c.tot_base * (c.num_semanas - 1)))::numeric, 2)
    END,
    CASE WHEN c.cuota_num <= c.pagos_hechos THEN 'PAGADO' ELSE 'PENDIENTE' END,
    CASE WHEN c.cuota_num <= c.pagos_hechos THEN NOW() ELSE NULL END,
    CASE
      WHEN c.cuota_num <= c.pagos_hechos THEN
        CASE
          WHEN c.cuota_num < c.num_semanas THEN c.tot_base
          ELSE ROUND((c.total_pagar - (c.tot_base * (c.num_semanas - 1)))::numeric, 2)
        END
      ELSE 0
    END,
    'Cuota ' || c.cuota_num || ' de ' || c.num_semanas,
    NOW()
  FROM calc c
  RETURNING prestamo_id
)
SELECT
  (SELECT count(*) FROM excel_data) AS filas_excel,
  (SELECT count(*) FROM matches) AS filas_matcheadas,
  (SELECT count(*) FROM updated) AS prestamos_actualizados,
  (SELECT count(*) FROM deleted) AS cuotas_eliminadas,
  (SELECT count(*) FROM ins) AS cuotas_insertadas,
  (SELECT count(*) FROM excel_data) - (SELECT count(*) FROM matches) AS no_encontradas;

COMMIT;
"""


def write_report(rows):
    # Reporte simple para trazabilidad local
    import csv
    with open(OUTPUT_REPORT, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow([
            "excel_row", "nombre_completo", "fecha_inicio", "monto_solicitado",
            "num_semanas", "pagos_hechos", "pagos_pendientes", "pendiente", "status"
        ])
        for row in rows:
            w.writerow([
                row["excel_row"], row["nombre_completo"], row["fecha_inicio"],
                row["monto_solicitado"], row["num_semanas"], row["pagos_hechos"],
                row["pagos_pendientes"], row["pendiente"], row["status"]
            ])


def main():
    rows = parse_rows()
    if not rows:
        raise RuntimeError("No se detectaron filas válidas en el Excel")

    sql = build_sql(rows)
    with open(OUTPUT_SQL, "w", encoding="utf-8") as f:
        f.write(sql)

    write_report(rows)

    print(f"OK: SQL generado -> {OUTPUT_SQL}")
    print(f"OK: Reporte parse -> {OUTPUT_REPORT}")
    print(f"Filas procesadas: {len(rows)}")


if __name__ == "__main__":
    main()
