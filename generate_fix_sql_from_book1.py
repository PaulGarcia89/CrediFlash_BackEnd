#!/usr/bin/env python3
import datetime as dt
import os
from decimal import Decimal, InvalidOperation

import openpyxl


EXCEL_PATH = os.getenv("EXCEL_PATH", "/Users/paulgarcia/Downloads/Book1.xlsx")
OUTPUT_SQL = os.getenv("OUTPUT_SQL", "fix_prestamos_from_book1.sql")


def to_decimal(value, default=Decimal("0")):
    if value is None:
        return default
    if isinstance(value, (int, float, Decimal)):
        try:
            return Decimal(str(value))
        except InvalidOperation:
            return default
    txt = str(value).strip()
    if txt in {"", "-", "None", "null"}:
        return default
    txt = txt.replace(",", "")
    txt = txt.replace("$", "")
    txt = txt.replace("%", "")
    try:
        return Decimal(txt)
    except InvalidOperation:
        return default


def to_int(value, default=0):
    return int(to_decimal(value, Decimal(default)))


def to_date(value):
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if value is None:
        return None
    txt = str(value).strip()
    if not txt or txt == "-":
        return None
    for fmt in ("%d-%b-%y", "%d-%b-%Y", "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return dt.datetime.strptime(txt, fmt).date()
        except ValueError:
            continue
    return None


def clean_text(value):
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def esc_sql(text):
    return text.replace("'", "''")


def parse_rows():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["Sheet1"]

    rows = []
    seq_by_key = {}

    for r in range(2, ws.max_row + 1):
        nombre = clean_text(ws.cell(r, 4).value)
        if not nombre:
            continue

        fecha_inicio = to_date(ws.cell(r, 1).value)
        if not fecha_inicio:
            continue

        key = (nombre.upper(), fecha_inicio.isoformat())
        seq_by_key[key] = seq_by_key.get(key, 0) + 1
        seq_excel = seq_by_key[key]

        monto_solicitado = to_decimal(ws.cell(r, 5).value)
        interes = to_int(ws.cell(r, 6).value, 0)
        modalidad = clean_text(ws.cell(r, 7).value) or "SEMANAL"
        num_semanas = to_int(ws.cell(r, 8).value, 0)
        num_dias = to_int(ws.cell(r, 9).value, 0)
        fecha_vencimiento = to_date(ws.cell(r, 10).value)
        total_pagar = to_decimal(ws.cell(r, 11).value)
        ganancias = to_decimal(ws.cell(r, 12).value)
        pagos_semanales = to_decimal(ws.cell(r, 13).value)
        pagos_hechos = to_decimal(ws.cell(r, 26).value)
        pagos_pendientes = to_decimal(ws.cell(r, 27).value)
        pagado = to_decimal(ws.cell(r, 28).value)
        pendiente = to_decimal(ws.cell(r, 29).value)
        status = clean_text(ws.cell(r, 30).value)

        if status.upper() in {"NO DEBE NADA", "PAGADO"}:
            status = "PAGADO"

        rows.append(
            {
                "seq_excel": seq_excel,
                "nombre_completo": nombre,
                "fecha_inicio": fecha_inicio.isoformat(),
                "monto_solicitado": monto_solicitado,
                "interes": interes,
                "modalidad": modalidad.upper(),
                "num_semanas": num_semanas,
                "num_dias": num_dias,
                "fecha_vencimiento": fecha_vencimiento.isoformat() if fecha_vencimiento else None,
                "total_pagar": total_pagar,
                "ganancias": ganancias,
                "pagos_semanales": pagos_semanales,
                "pagos_hechos": pagos_hechos,
                "pagos_pendientes": pagos_pendientes,
                "pagado": pagado,
                "pendiente": pendiente,
                "status": status,
            }
        )

    return rows


def build_sql(rows):
    values_lines = []
    for row in rows:
        fecha_vto_sql = f"'{row['fecha_vencimiento']}'::date" if row["fecha_vencimiento"] else "NULL"
        status_sql = f"'{esc_sql(row['status'])}'" if row["status"] else "NULL"
        values_lines.append(
            "("
            f"'{esc_sql(row['nombre_completo'])}', "
            f"'{row['fecha_inicio']}'::date, "
            f"{row['seq_excel']}, "
            f"{row['monto_solicitado']}, "
            f"{row['interes']}, "
            f"'{esc_sql(row['modalidad'])}', "
            f"{row['num_semanas']}, "
            f"{row['num_dias']}, "
            f"{fecha_vto_sql}, "
            f"{row['total_pagar']}, "
            f"{row['ganancias']}, "
            f"{row['pagos_semanales']}, "
            f"{row['pagos_hechos']}, "
            f"{row['pagos_pendientes']}, "
            f"{row['pagado']}, "
            f"{row['pendiente']}, "
            f"{status_sql}"
            ")"
        )

    values_block = ",\n    ".join(values_lines)

    return f"""-- Generado automáticamente desde {EXCEL_PATH}
BEGIN;

WITH excel_data AS (
  SELECT * FROM (
    VALUES
    {values_block}
  ) AS v(
    nombre_completo,
    fecha_inicio,
    seq_excel,
    monto_solicitado,
    interes,
    modalidad,
    num_semanas,
    num_dias,
    fecha_vencimiento,
    total_pagar,
    ganancias,
    pagos_semanales,
    pagos_hechos,
    pagos_pendientes,
    pagado,
    pendiente,
    status
  )
),
db_rows AS (
  SELECT
    p.id,
    upper(trim(regexp_replace(p.nombre_completo, '\\\\s+', ' ', 'g'))) AS nombre_norm,
    p.fecha_inicio::date AS fecha_inicio,
    row_number() OVER (
      PARTITION BY upper(trim(regexp_replace(p.nombre_completo, '\\\\s+', ' ', 'g'))), p.fecha_inicio::date
      ORDER BY p.id
    ) AS seq_db
  FROM public.prestamos p
),
xls_norm AS (
  SELECT
    upper(trim(regexp_replace(e.nombre_completo, '\\\\s+', ' ', 'g'))) AS nombre_norm,
    e.*
  FROM excel_data e
),
matches AS (
  SELECT
    d.id,
    x.*
  FROM db_rows d
  JOIN xls_norm x
    ON x.nombre_norm = d.nombre_norm
   AND x.fecha_inicio = d.fecha_inicio
   AND x.seq_excel = d.seq_db
),
updated AS (
  UPDATE public.prestamos p
  SET
    monto_solicitado = m.monto_solicitado,
    interes = m.interes,
    modalidad = m.modalidad,
    num_semanas = m.num_semanas,
    num_dias = m.num_dias,
    fecha_vencimiento = m.fecha_vencimiento,
    total_pagar = m.total_pagar,
    ganancias = m.ganancias,
    pagos_semanales = m.pagos_semanales,
    pagos_hechos = m.pagos_hechos,
    pagos_pendientes = m.pagos_pendientes,
    pagado = m.pagado,
    pendiente = m.pendiente,
    status = m.status
  FROM matches m
  WHERE p.id = m.id
  RETURNING p.id
)
SELECT count(*) AS filas_actualizadas FROM updated;

COMMIT;
"""


def main():
    rows = parse_rows()
    if not rows:
        raise RuntimeError("No se encontraron filas válidas en el Excel.")

    sql = build_sql(rows)
    with open(OUTPUT_SQL, "w", encoding="utf-8") as f:
        f.write(sql)

    print(f"OK: script SQL generado -> {OUTPUT_SQL}")
    print(f"Filas procesadas desde Excel: {len(rows)}")


if __name__ == "__main__":
    main()
