import os
import re
import csv
import uuid
import hashlib
from datetime import datetime, timedelta, date
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation

import psycopg2
from psycopg2.extras import execute_values
from openpyxl import load_workbook


# =========================
# Config
# =========================
DATABASE_URL = os.environ["DATABASE_URL"]
EXCEL_PATH = os.environ["EXCEL_PATH"]
DOCS_DIR = os.environ.get("DOCS_DIR", "").strip() or None
DRY_RUN = os.environ.get("DRY_RUN", "false").lower() == "true"

SHEET_CLIENTES = "REGISTRO CLIENTES"
SHEET_PRESTAMOS = "CONTROL PRESTAMOS"

OUT_ERRORS_CSV = "reporte_errores_importacion.csv"
ORIGEN_TAG = "CARGA_EXCEL_MARZO_2026"

# si quieres forzar modalidad por defecto cuando venga vacía/inválida
DEFAULT_MODALIDAD = "SEMANAL"


# =========================
# Helpers
# =========================
def d(val, default="0.00"):
    try:
        if val is None or str(val).strip() == "":
            return Decimal(default).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        return Decimal(str(val)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError):
        return Decimal(default).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

def d4(val, default="0.0000"):
    try:
        if val is None or str(val).strip() == "":
            return Decimal(default).quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)
        return Decimal(str(val)).quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError):
        return Decimal(default).quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)

def clean(s):
    return str(s or "").strip()

def upper(s):
    return clean(s).upper()

SPANISH_MONTHS = {
    "ENE": 1,
    "FEB": 2,
    "MAR": 3,
    "ABR": 4,
    "MAY": 5,
    "JUN": 6,
    "JUL": 7,
    "AGO": 8,
    "SEP": 9,
    "SET": 9,
    "OCT": 10,
    "NOV": 11,
    "DIC": 12,
}


def as_date(v, fallback_year=None):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    text = clean(v)
    if not text:
        return None

    normalized = (
        text.upper()
        .replace(".", "")
        .replace("º", "")
        .replace("ª", "")
    )

    match = re.match(r"^(\d{1,2})-([A-ZÑ]+)(?:-(\d{2,4}))?$", normalized)
    if match:
        day = int(match.group(1))
        month_key = match.group(2)[:3]
        month = SPANISH_MONTHS.get(month_key)
        if not month:
            return None
        year_text = match.group(3)
        if year_text is None and fallback_year is not None:
            year = int(fallback_year)
        elif year_text is not None:
            year = int(year_text)
            if year < 100:
                year += 2000
        else:
            return None
        try:
            return date(year, month, day)
        except ValueError:
            return None

    for fmt in ("%Y-%m-%d", "%d-%b-%y", "%d-%b-%Y", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    return None

def to_int(v, default=0):
    try:
        if v is None or str(v).strip() == "":
            return default
        return int(float(v))
    except Exception:
        return default

def split_name(full_name):
    parts = clean(full_name).split()
    if not parts:
        return ("N/A", "N/A")
    if len(parts) == 1:
        return (parts[0], "N/A")
    return (parts[0], " ".join(parts[1:]))

def normalize_email(email_raw):
    e = clean(email_raw).lower()
    if not e:
        return None
    # email simple, tolerante
    if "@" not in e or "." not in e.split("@")[-1]:
        return None
    return e[:100]

def normalize_phone(phone_raw):
    p = clean(phone_raw)
    if not p:
        return None
    return p[:20]

def parse_interes_to_tasa(interes_excel):
    # Excel viene muchas veces como 0.14 o 14
    if interes_excel is None or clean(interes_excel) == "":
        return d4("0.1200")
    x = d4(interes_excel)
    if x > Decimal("1"):
        x = (x / Decimal("100")).quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)
    if x < Decimal("0"):
        x = d4("0.1200")
    return x

def normalize_modalidad(m):
    v = upper(m)
    if v in {"SEMANAL", "QUINCENAL", "MENSUAL"}:
        return v
    return DEFAULT_MODALIDAD

def extract_pending_from_status(status):
    txt = upper(status)
    m = re.search(r"LE QUEDAN\s+(\d+)\s+PAGOS", txt)
    if m:
        return int(m.group(1))
    if txt == "NO DEBE NADA":
        return 0
    return None

def safe_status(estatus, pagos_pendientes):
    txt = clean(estatus)
    if txt:
        return txt[:100]
    if pagos_pendientes == 0:
        return "NO DEBE NADA"
    return f"LE QUEDAN {pagos_pendientes} PAGOS POR PAGAR"

def is_checked_cell(value):
    if value is True:
        return True
    if value is False or value is None:
        return False

    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value)) > 0

    text = upper(value).replace(".", "").strip()

    return text in {
        "X", "SI", "SÍ", "TRUE", "VERDADERO", "1",
        "✓", "✔", "☑", "☒", "CHEK", "CHECK", "CHECKED"
    }

def document_candidates_for_name(nombre):
    # búsqueda flexible de pdf por nombre
    if not DOCS_DIR or not os.path.isdir(DOCS_DIR):
        return []
    base = upper(nombre)
    tokens = re.sub(r"[^A-Z0-9 ]", " ", base).split()
    out = []
    for fname in os.listdir(DOCS_DIR):
        if not fname.lower().endswith(".pdf"):
            continue
        fup = upper(fname)
        score = sum(1 for t in tokens if t in fup)
        if score >= max(1, min(2, len(tokens))):
            out.append((score, fname))
    out.sort(reverse=True)
    return [x[1] for x in out[:3]]

def rel_doc_path(filename):
    # tu backend usa rutas tipo uploads/solicitudes/archivo.pdf
    return f"uploads/solicitudes/{filename}".replace("\\", "/")

def get_sheet_by_candidates(wb, candidates, required=False):
    available = {name.upper(): name for name in wb.sheetnames}
    for candidate in candidates:
        key = upper(candidate)
        if key in available:
            return wb[available[key]]
    if required:
        raise KeyError(f"Ninguna hoja encontrada entre: {candidates}. Hojas disponibles: {wb.sheetnames}")
    return None


# =========================
# DB utilities
# =========================
def ensure_columns(cur):
    cur.execute("ALTER TABLE public.clientes ADD COLUMN IF NOT EXISTS origen character varying(40)")
    cur.execute("ALTER TABLE public.solicitudes ADD COLUMN IF NOT EXISTS origen character varying(40)")
    cur.execute("ALTER TABLE public.solicitud_documentos ADD COLUMN IF NOT EXISTS tipo_documento character varying(30)")
    cur.execute("ALTER TABLE public.prestamos ADD COLUMN IF NOT EXISTS modalidad character varying(30)")
    cur.execute("ALTER TABLE public.prestamos ADD COLUMN IF NOT EXISTS fecha_aprobacion timestamp without time zone NULL")

def reset_business_tables(cur):
    cur.execute("""
        TRUNCATE TABLE
          public.cuotas,
          public.prestamos,
          public.solicitud_documentos,
          public.solicitudes,
          public.clientes_email_verificaciones,
          public.clientes
        RESTART IDENTITY CASCADE
    """)

def upsert_default_modelo_aprobacion(cur):
    cur.execute("""
        SELECT id FROM public.modelos_aprobacion
        WHERE nombre = 'Modelo Cliente Antiguo'
        LIMIT 1
    """)
    row = cur.fetchone()
    if row:
        return row[0]

    new_id = str(uuid.uuid4())
    cur.execute("""
        INSERT INTO public.modelos_aprobacion (id, nombre, reglas, puntaje_minimo, activo, creado_en)
        VALUES (%s, %s, %s::jsonb, %s, %s, NOW())
    """, (new_id, "Modelo Cliente Antiguo", "{}", 0, True))
    return new_id


# =========================
# Parsing Excel
# =========================
def load_client_rows(wb):
    ws = get_sheet_by_candidates(wb, [SHEET_CLIENTES], required=False)
    if ws is None:
        print(f"ℹ️ Hoja '{SHEET_CLIENTES}' no encontrada. Se crearán clientes desde la hoja de préstamos.")
        return []
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        nombre = clean(r[2])
        if not nombre or upper(nombre) in {"NOMBRE", "TOTAL"}:
            continue
        rows.append({
            "nombre_full": nombre,
            "telefono": normalize_phone(r[3]),
            "email": normalize_email(r[4]),
            "contacto_nombre": clean(r[5])[:100] or None,
            "contacto_telefono": normalize_phone(r[6]),
            "referido_por": clean(r[7])[:150] or None,
            "calificacion": clean(r[8])[:200] or None
        })
    return rows

def load_loan_rows(wb):
    ws = get_sheet_by_candidates(wb, [SHEET_PRESTAMOS, "Sheet1", "SHEET1", "CONTROL PRESTAMOS"], required=True)
    rows = []
    for idx, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        nombre = clean(r[3])
        if not nombre or upper(nombre) == "NOMBRE":
            continue

        fecha_inicio = as_date(r[0], fallback_year=to_int(r[2], datetime.now().year)) or datetime.now().date()
        modalidad = normalize_modalidad(r[6])
        semanas = to_int(r[7], 0)
        dias = to_int(r[8], max(0, semanas * 7))
        fecha_venc = as_date(r[9], fallback_year=fecha_inicio.year) or (fecha_inicio + timedelta(days=max(1, semanas) * 7))

        monto = d(r[4])
        tasa = parse_interes_to_tasa(r[5])
        total = d(r[10], default=str(monto))
        ganancias = d(r[11], default="0.00")
        pago_semanal = d(r[12], default="0.00")

        # Resumen real de esta hoja: columnas 14 a 18
        pagos_hechos_col = to_int(r[13], 0)
        pagos_pend_col = to_int(r[14], 0)
        pagado_col = d(r[15], default="0.00")
        balance_col = d(r[16], default="0.00")
        estatus_col = clean(r[17])

        # En esta hoja no hay checkboxes confiables, así que usamos el valor numérico
        pagos_hechos_checks = pagos_hechos_col

        if semanas <= 0:
            # fallback robusto
            if pago_semanal > 0:
                semanas = int((total / pago_semanal).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
            if semanas <= 0:
                semanas = max(1, pagos_hechos)

        # Priorizamos columnas resumen si vienen informadas
        resumen_informado = (
            clean(r[13]) not in {"", "-"} or
            clean(r[14]) not in {"", "-"} or
            clean(r[17]) not in {"", "-"}
        )

        if resumen_informado:
            pending_from_status = extract_pending_from_status(estatus_col)
            if pending_from_status is not None:
                pagos_pend = pending_from_status
                pagos_hechos = max(0, semanas - pagos_pend)
            else:
                pagos_hechos = max(0, pagos_hechos_col)
                pagos_pend = max(0, pagos_pend_col) if pagos_pend_col > 0 else max(0, semanas - pagos_hechos)
        else:
            pagos_hechos = max(0, pagos_hechos_checks)
            pagos_pend = max(0, semanas - pagos_hechos)

        if pagos_hechos > semanas:
            pagos_hechos = semanas
        if pagos_pend > semanas:
            pagos_pend = semanas

        if pago_semanal <= 0 and semanas > 0:
            pago_semanal = (total / Decimal(semanas)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

        if total <= 0:
            total = (monto + (monto * tasa * Decimal(semanas))).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

        if ganancias <= 0:
            ganancias = (total - monto).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

        if resumen_informado and pagado_col > 0:
            pagado = pagado_col
        else:
            pagado = (pago_semanal * Decimal(pagos_hechos)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            if pagado > total:
                pagado = total

        if resumen_informado and balance_col > 0:
            balance = balance_col
        else:
            balance = max(Decimal("0.00"), (total - pagado)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

        estatus = "NO DEBE NADA" if pagos_pend == 0 else f"LE QUEDAN {pagos_pend} PAGOS POR PAGAR"

        rows.append({
            "excel_row": idx,
            "nombre_full": nombre,
            "fecha_inicio": fecha_inicio,
            "modalidad": modalidad,
            "num_semanas": semanas,
            "num_dias": dias,
            "fecha_vencimiento": fecha_venc,
            "monto_solicitado": monto,
            "tasa_variable": tasa,
            "total_pagar": total,
            "ganancias": ganancias,
            "pagos_semanales": pago_semanal,
            "pagos_hechos": pagos_hechos,
            "pagos_pendientes": pagos_pend,
            "pagado": pagado,
            "pendiente": balance,
            "status": estatus
        })

    return rows


# =========================
# Main import
# =========================
def main():
    wb = load_workbook(EXCEL_PATH, data_only=True)

    client_rows = load_client_rows(wb)
    loan_rows = load_loan_rows(wb)

    errors = []
    created = {"clientes": 0, "solicitudes": 0, "prestamos": 0, "cuotas": 0, "documentos": 0}

    conn = psycopg2.connect(DATABASE_URL)
    conn.autocommit = False
    cur = conn.cursor()

    try:
        ensure_columns(cur)
        if not DRY_RUN:
            reset_business_tables(cur)

        modelo_aprobacion_id = upsert_default_modelo_aprobacion(cur)

        # mapa cliente por nombre
        clients_by_name = {}

        # 1) clientes
        for c in client_rows:
            try:
                cid = str(uuid.uuid4())
                n, a = split_name(c["nombre_full"])
                observ = f"{ORIGEN_TAG}" + (f" | CALIFICACION:{c['calificacion']}" if c["calificacion"] else "")
                referido_flag = bool(c["referido_por"])

                if not DRY_RUN:
                    cur.execute("""
                        INSERT INTO public.clientes (
                          id, fecha_registro, nombre, apellido, telefono, email, direccion,
                          nombre_contacto, apellido_contacto, telefono_contacto, email_contacto, direccion_contacto,
                          es_referido, referido_por, monto_referido, descuentos_referido_disponibles, descuentos_referido_aplicados,
                          estado, observaciones, origen
                        ) VALUES (
                          %s, NOW(), %s, %s, %s, %s, NULL,
                          %s, NULL, %s, NULL, NULL,
                          %s, %s, 0, 0, 0,
                          'ACTIVO', %s, 'PUBLIC_FORM'
                        )
                    """, (
                        cid, n[:100], a[:100], c["telefono"], c["email"],
                        c["contacto_nombre"], c["contacto_telefono"],
                        referido_flag, c["referido_por"], observ[:500]
                    ))

                clients_by_name[upper(c["nombre_full"])] = cid
                created["clientes"] += 1
            except Exception as e:
                errors.append([SHEET_CLIENTES, "-", c["nombre_full"], f"ERROR_CLIENTE: {e}"])

        # 2) solicitudes + prestamos + cuotas + documentos
        for p in loan_rows:
            try:
                key = upper(p["nombre_full"])
                cliente_id = clients_by_name.get(key)

                if not cliente_id:
                    # crear cliente fallback
                    cid = str(uuid.uuid4())
                    n, a = split_name(p["nombre_full"])
                    if not DRY_RUN:
                        cur.execute("""
                            INSERT INTO public.clientes (
                              id, fecha_registro, nombre, apellido, estado, observaciones, origen
                            ) VALUES (%s, NOW(), %s, %s, 'ACTIVO', %s, 'PUBLIC_FORM')
                        """, (cid, n[:100], a[:100], f"{ORIGEN_TAG} | CREADO_DESDE_PRESTAMO"))
                    cliente_id = cid
                    clients_by_name[key] = cid
                    created["clientes"] += 1

                solicitud_id = str(uuid.uuid4())
                prestamo_id = str(uuid.uuid4())

                interes_pct_int = int((p["tasa_variable"] * Decimal("100")).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
                fecha_ini_dt = datetime.combine(p["fecha_inicio"], datetime.min.time())
                fecha_venc_dt = datetime.combine(p["fecha_vencimiento"], datetime.min.time())

                if not DRY_RUN:
                    cur.execute("""
                        INSERT INTO public.solicitudes (
                          id, cliente_id, analista_id, modelo_aprobacion_id, modelo_calificacion,
                          modalidad, tasa_base, monto_solicitado, plazo_semanas, tasa_variable,
                          estado, creado_en, fecha_aprobacion, destino, origen
                        ) VALUES (
                          %s, %s, NULL, %s, 'CLIENTE_ANTIGUO',
                          %s, %s, %s, %s, %s,
                          'APROBADO', %s, %s, 'CARGA_EXCEL', 'PUBLIC_FORM'
                        )
                    """, (
                        solicitud_id, cliente_id, modelo_aprobacion_id,
                        p["modalidad"], p["tasa_variable"], p["monto_solicitado"], p["num_semanas"], p["tasa_variable"],
                        fecha_ini_dt, fecha_ini_dt
                    ))

                    cur.execute("""
                        INSERT INTO public.prestamos (
                          id, solicitud_id, contrato, recordatorio_whatsapp_modo,
                          fecha_inicio, mes, anio, nombre_completo, monto_solicitado, interes, modalidad,
                          num_semanas, num_dias, fecha_vencimiento, fecha_aprobacion,
                          total_pagar, ganancias, pagos_semanales, pagos_hechos, pagos_pendientes,
                          pagado, pendiente, status, ganancia_diaria, reserva, refinanciado, perdida,
                          caso_especial, oferta, proyeccion_mes, anio_vencimiento
                        ) VALUES (
                          %s, %s, NULL, 'AUTO',
                          %s, %s, %s, %s, %s, %s, %s,
                          %s, %s, %s, %s,
                          %s, %s, %s, %s, %s,
                          %s, %s, %s, 0, 0, 0, 0,
                          NULL, 0, NULL, NULL
                        )
                    """, (
                        prestamo_id, solicitud_id,
                        fecha_ini_dt, p["fecha_inicio"].strftime("%b"), str(p["fecha_inicio"].year),
                        p["nombre_full"][:200], p["monto_solicitado"], interes_pct_int, p["modalidad"],
                        p["num_semanas"], p["num_dias"], fecha_venc_dt, fecha_ini_dt,
                        p["total_pagar"], p["ganancias"], p["pagos_semanales"], p["pagos_hechos"], p["pagos_pendientes"],
                        p["pagado"], p["pendiente"], p["status"][:100]
                    ))

                created["solicitudes"] += 1
                created["prestamos"] += 1

                # cuotas según semanas, y si status dice LE QUEDAN X PAGOS... respeta pendientes
                total_weeks = max(1, p["num_semanas"])
                remaining_paid_pool = p["pagado"]
                interes_cuota = (p["ganancias"] / Decimal(total_weeks)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

                cuotas_rows = []
                for i in range(1, total_weeks + 1):
                    cuota_id = str(uuid.uuid4())
                    venc = p["fecha_inicio"] + timedelta(days=7 * i)

                    monto_total = p["pagos_semanales"]
                    monto_interes = interes_cuota
                    monto_capital = (monto_total - monto_interes).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                    if monto_capital < 0:
                        monto_capital = Decimal("0.00")

                    if remaining_paid_pool >= monto_total:
                        monto_pagado = monto_total
                        estado = "PAGADO"
                        fecha_pago = datetime.combine(venc, datetime.min.time())
                    elif remaining_paid_pool > 0:
                        monto_pagado = remaining_paid_pool
                        estado = "PENDIENTE"
                        fecha_pago = None
                    else:
                        monto_pagado = Decimal("0.00")
                        estado = "PENDIENTE"
                        fecha_pago = None

                    remaining_paid_pool = (remaining_paid_pool - monto_pagado).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                    obs = f"Cuota {i} de {total_weeks}"

                    cuotas_rows.append((
                        cuota_id, prestamo_id, venc, monto_capital, monto_interes, monto_total,
                        estado, fecha_pago, monto_pagado, Decimal("0.00"), Decimal("0.00"), None, obs
                    ))

                if not DRY_RUN and cuotas_rows:
                    execute_values(cur, """
                        INSERT INTO public.cuotas (
                          id, prestamo_id, fecha_vencimiento, monto_capital, monto_interes, monto_total,
                          estado, fecha_pago, monto_pagado, monto_fee_acumulado, monto_penalizacion_acumulada,
                          motivo_fee, observaciones, created_at
                        ) VALUES %s
                    """, [r + (datetime.now(),) for r in cuotas_rows])

                created["cuotas"] += len(cuotas_rows)

                # documentos opcionales por matching de nombre en DOCS_DIR
                if DOCS_DIR:
                    candidates = document_candidates_for_name(p["nombre_full"])
                    if candidates and not DRY_RUN:
                        for j, fname in enumerate(candidates):
                            doc_id = str(uuid.uuid4())
                            tipo_doc = "ID" if j == 0 else "ESTADO_CUENTA"
                            full_path = os.path.join(DOCS_DIR, fname)
                            try:
                                size = os.path.getsize(full_path)
                            except Exception:
                                size = 0
                            cur.execute("""
                                INSERT INTO public.solicitud_documentos (
                                  id, solicitud_id, prestamo_id, nombre_original, nombre_archivo, mime_type,
                                  size_bytes, tipo_documento, ruta, creado_en
                                ) VALUES (
                                  %s, %s, %s, %s, %s, 'application/pdf',
                                  %s, %s, %s, NOW()
                                )
                            """, (
                                doc_id, solicitud_id, prestamo_id, fname[:255], fname[:255],
                                size, tipo_doc, rel_doc_path(fname)
                            ))
                            created["documentos"] += 1

            except Exception as e:
                errors.append([SHEET_PRESTAMOS, p["excel_row"], p["nombre_full"], f"ERROR_PRESTAMO: {e}"])

        if DRY_RUN:
            conn.rollback()
            print("🧪 DRY_RUN=true -> ROLLBACK ejecutado (sin cambios en DB)")
        else:
            conn.commit()
            print("✅ COMMIT OK")

    except Exception as e:
        conn.rollback()
        print(f"❌ ERROR GENERAL, rollback aplicado: {e}")
        raise
    finally:
        cur.close()
        conn.close()

    # reporte de errores
    with open(OUT_ERRORS_CSV, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["hoja", "fila_excel", "cliente", "error"])
        w.writerows(errors)

    print("\n=== RESUMEN ===")
    for k, v in created.items():
        print(f"{k}: {v}")
    print(f"errors: {len(errors)}")
    print(f"archivo errores: {OUT_ERRORS_CSV}")


if __name__ == "__main__":
    main()
